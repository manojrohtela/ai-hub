import base64
import json
import re
import sqlite3
from pathlib import Path

from groq import Groq

from .config import get_settings

DB_PATH = Path(__file__).parent / "scores.db"


def _conn() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def _init_db() -> None:
    with _conn() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS matches (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS scores (
                player_name TEXT    NOT NULL,
                match_id    INTEGER NOT NULL REFERENCES matches(id),
                points      INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (player_name, match_id)
            )
        """)


class ScoreKeeperService:
    def __init__(self):
        self.groq = Groq(api_key=get_settings().groq_api_key)
        _init_db()

    # ------------------------------------------------------------------ #

    def parse_image(self, image_bytes: bytes, content_type: str) -> dict[str, int]:
        b64 = base64.b64encode(image_bytes).decode()
        completion = self.groq.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:{content_type};base64,{b64}"},
                    },
                    {
                        "type": "text",
                        "text": (
                            "Extract every player name and their numeric score/points "
                            "from this image.\n"
                            "Return ONLY valid JSON — no markdown, no explanation:\n"
                            '{"players": [{"name": "...", "points": 42}, ...]}\n'
                            "Keep names exactly as shown (Hindi or any script is fine)."
                        ),
                    },
                ],
            }],
            max_tokens=1024,
        )
        raw = re.sub(r"```(?:json)?", "", completion.choices[0].message.content or "").strip()
        try:
            result = json.loads(raw)
        except Exception:
            m = re.search(r"\{.*\}", raw, re.DOTALL)
            result = json.loads(m.group()) if m else {}

        return {
            p["name"].strip(): int(p["points"])
            for p in result.get("players", [])
            if p.get("name") and str(p.get("points", "")).lstrip("-").isdigit()
        }

    def add_match(self, player_scores: dict[str, int]) -> tuple[str, int]:
        with _conn() as con:
            count = con.execute("SELECT COUNT(*) FROM matches").fetchone()[0]
            match_number = count + 1
            match_name = f"Match {match_number}"
            cur = con.execute("INSERT INTO matches (name) VALUES (?)", (match_name,))
            match_id = cur.lastrowid
            for name, points in player_scores.items():
                con.execute(
                    "INSERT INTO scores (player_name, match_id, points) VALUES (?,?,?)"
                    " ON CONFLICT(player_name, match_id) DO UPDATE SET points=excluded.points",
                    (name, match_id, points),
                )
        return match_name, match_number

    def get_standings(self) -> dict:
        with _conn() as con:
            matches = [dict(r) for r in con.execute("SELECT id, name FROM matches ORDER BY id")]
            rows = con.execute("SELECT player_name, match_id, points FROM scores").fetchall()

        players: dict[str, dict[int, int]] = {}
        for row in rows:
            players.setdefault(row["player_name"], {})[row["match_id"]] = row["points"]

        standings = []
        for name, score_map in players.items():
            total = sum(score_map.values())
            match_points = {m["name"]: score_map.get(m["id"], 0) for m in matches}
            standings.append({"name": name, "total": total, "matches": match_points})

        standings.sort(key=lambda x: x["total"], reverse=True)
        for i, s in enumerate(standings):
            s["rank"] = i + 1

        return {"players": standings, "match_headers": [m["name"] for m in matches]}

    def export_xlsx(self) -> bytes:
        import io
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        data = self.get_standings()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scores"

        headers = ["Rank", "Player"] + data["match_headers"] + ["Total"]
        hfill = PatternFill("solid", fgColor="1e293b")
        hfont = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal="center")

        for ri, p in enumerate(data["players"], 2):
            ws.cell(row=ri, column=1, value=p["rank"])
            ws.cell(row=ri, column=2, value=p["name"])
            for ci, mh in enumerate(data["match_headers"]):
                ws.cell(row=ri, column=3 + ci, value=p["matches"].get(mh, 0))
            ws.cell(row=ri, column=len(headers), value=p["total"])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def chat(self, question: str) -> str:
        data = self.get_standings()
        if not data["players"]:
            return "अभी तक कोई match data नहीं है। / No match data yet. Please upload a match image first."

        lines = ["=== Match Points Data ==="]
        for p in data["players"]:
            details = " | ".join(f"{m}: {v}" for m, v in p["matches"].items())
            lines.append(f"Rank {p['rank']}: {p['name']} — Total: {p['total']} | {details}")

        completion = self.groq.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a helpful points/score assistant for a game leaderboard. "
                        "Answer questions about player scores and match standings. "
                        "The user may ask in Hindi or English — always respond in the SAME language. "
                        "Be concise and friendly.\n\n" + "\n".join(lines)
                    ),
                },
                {"role": "user", "content": question},
            ],
            max_tokens=512,
        )
        return completion.choices[0].message.content or "Sorry, could not generate a response."


_service: ScoreKeeperService | None = None


def get_service() -> ScoreKeeperService:
    global _service
    if _service is None:
        _service = ScoreKeeperService()
    return _service
